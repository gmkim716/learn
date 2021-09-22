from openpyxl import load_workbook 
# wb = load_workbook("sample_formula.xlsx")
# ws =  wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws =  wb.active

# 수식 그대로 가져오고 있음
# evaluate 되지 않은 상태의 데이터는 None으로 표시
for row in ws.values:
    for cell in row:
        print(cell)

